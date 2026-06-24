#!/usr/bin/env python3
"""Builds the Greek electricity contract analysis workbook.
Reads bills_data.json (parsed bills) and market_offers.json (researched offers).
Produces example_model.xlsx with 8 sheets and
transparent formulas for the scenario/backtest engine.
"""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
bills = json.load(open(os.path.join(HERE, "example_bills.json")))
offers = json.load(open(os.path.join(HERE, "example_offers.json")))

META = bills["meta"]
SUPPLIES = bills["supplies"]
BILLS = bills["bills"]
OFFERS = offers["offers"]
REC = offers["recommendations"]   # {supply_id: offer_name}

# ---------- styling helpers ----------
HDR = Font(bold=True, color="FFFFFF", size=11)
HDRFILL = PatternFill("solid", fgColor="1F4E78")
SUBHDR = Font(bold=True, size=11, color="1F4E78")
TITLE = Font(bold=True, size=14, color="1F4E78")
BOLD = Font(bold=True)
GOOD = PatternFill("solid", fgColor="C6EFCE")
BADF = PatternFill("solid", fgColor="FFC7CE")
WARN = PatternFill("solid", fgColor="FFEB9C")
GREY = PatternFill("solid", fgColor="F2F2F2")
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right")
EUR = '#,##0.00\\ "€"'
PCT = '0.0%'
RATE = '0.00000'

wb = Workbook()

def style_header_row(ws, row, ncols, start=1):
    for c in range(start, start + ncols):
        cell = ws.cell(row=row, column=c)
        cell.font = HDR; cell.fill = HDRFILL; cell.alignment = WRAP; cell.border = BORDER

def autosize(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

# =====================================================================
# SHEET 1: Bills_Raw
# =====================================================================
ws = wb.active
ws.title = "Bills_Raw"
ws["A1"] = "Bills_Raw (SYNTHETIC DEMO DATA) — one row per bill, all extracted fields (source of truth from the 8 PDFs)"
ws["A1"].font = TITLE
cols = ["File","Supply","Provider","Product","Type","Period start","Period end","Days",
        "Issue date","kWh total","kWh day","kWh night",
        "Energy rate (day)","Energy rate (night)","Adj. clause rate","Loyalty disc %","EN.A disc €","Paygio €/mo",
        "A: Supply chgs €","B: Regulated €","Γ: Supplementary €","EFK €","Special 5‰ €","Late interest €",
        "E: Municipal €","ΣΤ: ERT €","VAT €","VAT base €","Enanti credit €",
        "Prev. unpaid €","Current-period total €","Final payable €"]
hr = 3
for i, c in enumerate(cols, 1):
    ws.cell(row=hr, column=i, value=c)
style_header_row(ws, hr, len(cols))
r = hr + 1
for b in BILLS:
    vals = [b["file"], b["supply_id"], b["provider"], b["product"], b.get("type",""),
            b["period_start"], b["period_end"], b["days"], b["issue"],
            b["kwh_total"], b["kwh_day"], b["kwh_night"],
            b.get("energy_day_rate", b.get("energy_base_rate")), b.get("energy_night_rate"),
            b.get("adj_rate"), b.get("loyalty_disc_pct",0), b.get("ena_disc",0), b["fixed_paygio"],
            b["A_supply"], b["B_regulated"], b["G_supplementary"], b["efk"], b["special5permille"], b["late_interest"],
            b["E_municipal"], b["ST_ert"], b["vat"], b["vat_base"], b["enanti_credit"],
            b["prev_unpaid"], b["current_total"], b["final_payable"]]
    for i, v in enumerate(vals, 1):
        cell = ws.cell(row=r, column=i, value=v); cell.border = BORDER
        if i in (13,14,15): cell.number_format = RATE
        if i in (16,): cell.number_format = PCT
        if i in (17,18,19,20,21,22,23,24,25,26,27,28,29,30,31,32): cell.number_format = EUR
    r += 1
# totals
ws.cell(row=r, column=1, value="TOTAL").font = BOLD
for i in (19,20,21,25,26,27,31,32):
    col = get_column_letter(i)
    cell = ws.cell(row=r, column=i, value=f"=SUM({col}{hr+1}:{col}{r-1})")
    cell.font = BOLD; cell.number_format = EUR; cell.fill = GREY
ws.cell(row=r, column=10, value=f"=SUM(J{hr+1}:J{r-1})").font = BOLD
ws.freeze_panes = "B4"
autosize(ws, {"A":11,"B":7,"C":9,"D":20,"E":12,"F":12,"G":12,"H":6,"I":12,"J":9,"K":8,"L":9,
              "M":13,"N":13,"O":14,"P":12,"Q":11,"R":11,"S":15,"T":14,"U":17,"V":8,"W":11,"X":12,
              "Y":13,"Z":10,"AA":9,"AB":11,"AC":13,"AD":13,"AE":18,"AF":15})
note = ws.cell(row=r+2, column=1, value=(
    "Notes: (1) 5d2516a2 is a 66-day clearing bill whose payable (280.19) is NET of a 224.95€ 'enanti' (on-account) "
    "charge billed earlier. Its true economic period cost is reconstructed gross = 518.64€ (see Backtest). "
    "(2) VAT (6%) base = A + B + EFK. Municipal, ERT, special 5‰ and late interest are outside the VAT base. "
    "(3) EFK = 0.0022 €/kWh residential, 0.005 €/kWh business. (4) S4 (DEH) April municipal/ERT split estimated; aggregate exact."))
note.alignment = WRAP; note.font = Font(italic=True, size=9)
ws.merge_cells(start_row=r+2, start_column=1, end_row=r+4, end_column=12)

# =====================================================================
# SHEET 2: Supplies
# =====================================================================
ws = wb.create_sheet("Supplies")
ws["A1"] = "Supplies — one row per electricity supply/contract"
ws["A1"].font = TITLE
cols = ["Supply ID","Provider (now)","Supply number","Property / label","Customer type","Current product",
        "Tariff type","Power (kVA)","Phases","Register","Priority",
        "Bills count","Total kWh (sampled)","Total days (sampled)","Avg kWh/day","Annualised kWh est.","Notes"]
hr = 3
for i, c in enumerate(cols, 1):
    ws.cell(row=hr, column=i, value=c)
style_header_row(ws, hr, len(cols))
r = hr + 1
for s in SUPPLIES:
    sb = [b for b in BILLS if b["supply_id"] == s["id"]]
    tot_kwh = sum(b["kwh_total"] for b in sb)
    tot_days = sum(b["days"] for b in sb)
    avg = tot_kwh / tot_days if tot_days else 0
    ann = round(avg * 365)
    vals = [s["id"], s["provider"], s["supply_no"], s["property"], s["customer_type"], s["product"],
            s["tariff_type"], s["power_kva"], s["phases"], s["register"], s["priority"],
            len(sb), tot_kwh, tot_days, round(avg,2), ann, s.get("notes","")]
    for i, v in enumerate(vals, 1):
        cell = ws.cell(row=r, column=i, value=v); cell.border = BORDER; cell.alignment = WRAP
    if s["priority"] == "HIGH":
        ws.cell(row=r, column=11).fill = WARN; ws.cell(row=r, column=11).font = BOLD
    r += 1
autosize(ws, {"A":8,"B":11,"C":15,"D":34,"E":18,"F":18,"G":16,"H":9,"I":7,"J":12,"K":9,
              "L":7,"M":14,"N":12,"O":11,"P":15,"Q":40})
ws.freeze_panes = "A4"

# =====================================================================
# SHEET 3: Bill_Line_Items
# =====================================================================
ws = wb.create_sheet("Bill_Line_Items")
ws["A1"] = "Bill_Line_Items — detailed line items per bill"
ws["A1"].font = TITLE
cols = ["File","Supply","Category","Line item","Basis / calc","Amount €","Provider-dependent?"]
hr = 3
for i, c in enumerate(cols, 1):
    ws.cell(row=hr, column=i, value=c)
style_header_row(ws, hr, len(cols))
r = hr + 1

def line(ws, r, file, supp, cat, item, basis, amt, pd):
    vals = [file, supp, cat, item, basis, amt, pd]
    for i, v in enumerate(vals, 1):
        cell = ws.cell(row=r, column=i, value=v); cell.border = BORDER; cell.alignment = WRAP
        if i == 6: cell.number_format = EUR
        if i == 7 and v == "YES": cell.fill = WARN
    return r + 1

for b in BILLS:
    f = b["file"]; s = b["supply_id"]
    r = line(ws, r, f, s, "A Supply", "Paygio (fixed)", f"{b['fixed_paygio']:.2f}€ for {b['days']}d", b["fixed_paygio"], "YES")
    if b.get("energy_night_rate"):  # dual-register (day/night)
        r = line(ws, r, f, s, "A Supply", "Energy DAY (kanoniki)", f"{b['kwh_day']} kWh x {b['energy_day_rate']} €/kWh", round(b['kwh_day']*b['energy_day_rate'],2), "YES")
        r = line(ws, r, f, s, "A Supply", "Energy NIGHT (meiomeni)", f"{b['kwh_night']} kWh x {b['energy_night_rate']} €/kWh", round(b['kwh_night']*b['energy_night_rate'],2), "YES")
    else:  # single-register
        er = b.get("energy_base_rate") or b.get("energy_day_rate")
        r = line(ws, r, f, s, "A Supply", "Energy", f"{b['kwh_total']} kWh x {er} €/kWh", round(b['kwh_total']*er,2), "YES")
    if b.get("adj_rate"):
        r = line(ws, r, f, s, "A Supply", "Market adj. clause (Diakymansi)", f"{b['kwh_total']} kWh x {b['adj_rate']} €/kWh", round(b['kwh_total']*b['adj_rate'],2), "YES")
    if b.get("ena_disc"):
        r = line(ws, r, f, s, "A Supply", "Ekptosi EN.A", "product credit", b["ena_disc"], "YES")
    r = line(ws, r, f, s, "B Regulated", "ADMIE+DEDDIE+ETMEAR+YKO", "transmission+distribution+RES levy+PSO (regulated)", b["B_regulated"], "no")
    r = line(ws, r, f, s, "Taxes", "EFK (excise)", f"{b['efk']:.2f} (in VAT base)", b["efk"], "no")
    r = line(ws, r, f, s, "Taxes", "Special levy 5‰", "0.5% of energy value", b["special5permille"], "no")
    if b["late_interest"]: r = line(ws, r, f, s, "Taxes", "Late interest", "arrears-related", b["late_interest"], "no")
    if b["enanti_credit"]: r = line(ws, r, f, s, "Adjustment", "Axia Reumatos Enanti (credit)", "reverses prior on-account bill", b["enanti_credit"], "no")
    r = line(ws, r, f, s, "Municipal", "DT+DF+TAP (Dimos)", "per m2 / municipality", b["E_municipal"], "no")
    r = line(ws, r, f, s, "ERT", "ERT levy", "36€/yr prorated", b["ST_ert"], "no")
    r = line(ws, r, f, s, "VAT", "VAT 6%", f"6% x {b['vat_base']:.2f}", b["vat"], "no (follows A)")
    if b["prev_unpaid"]: r = line(ws, r, f, s, "Carry-over", "Previous unpaid", "prior bill balance", b["prev_unpaid"], "no")
    r += 1  # blank row between bills
autosize(ws, {"A":11,"B":7,"C":14,"D":30,"E":42,"F":11,"G":18})
ws.freeze_panes = "A4"

# =====================================================================
# SHEET 4: Fee_Model
# =====================================================================
ws = wb.create_sheet("Fee_Model")
ws["A1"] = "Fee_Model — how each charge is calculated and whether switching provider changes it"
ws["A1"].font = TITLE
cols = ["Charge","What it is","Who charges","Basis (depends on)","Changes with provider?","Avoidable/reducible?","How modelled in Excel"]
hr = 3
for i, c in enumerate(cols, 1):
    ws.cell(row=hr, column=i, value=c)
style_header_row(ws, hr, len(cols))
fee_rows = [
 ["Paygio (fixed supply fee)","Monthly standing charge of the supplier","Supplier (DEH/HERON/...)","Days (prorated)","YES","Yes – pick low/zero-paygio offer","paygio_monthly × days/30"],
 ["Energy charge","Competitive price for kWh consumed","Supplier","kWh (× day/night if dual)","YES","Yes – core of the comparison","kWh × rate (×(1-disc))"],
 ["Market adjustment clause (Diakymansi Kostous Agoras / ritra)","Pass-through of wholesale market cost on floating products","Supplier","kWh","YES","Yes – fixed-price product removes it","included in effective rate; 0 for fixed products"],
 ["Loyalty / quantity / consistency discount","Discount for staying / paying on time / direct debit","Supplier","% of energy or fixed","YES","Yes – maximise by qualifying","applied as % off energy or fixed €"],
 ["ADMIE (Transmission / Systima Metaforas)","Use of the high-voltage transmission grid","ADMIE (regulated)","kWh (+kVA capacity)","NO – same for all suppliers","No","kept as actual; pass-through unchanged"],
 ["DEDDIE (Distribution / Diktyo Dianomis)","Use of the low/medium-voltage distribution grid","DEDDIE (regulated)","kVA capacity × days + kWh","NO – same for all","No","kept as actual; pass-through unchanged"],
 ["ETMEAR","Levy financing renewables (RES special account)","State/regulated","kWh","NO – same for all","No","kept as actual (0.017 €/kWh observed)"],
 ["YKO (Public Service Obligations)","Funds subsidised tariffs/island supply","State/regulated","kWh (tiered; night band)","NO – same for all","No","kept as actual"],
 ["EFK (Excise / Eidikos Foros Katanalosis)","State excise on electricity","State","kWh","NO","No","0.0022 €/kWh resid., 0.005 business; in VAT base"],
 ["Special levy 5‰ (N.2093/92)","Stamp-type special duty","State","0.5% of energy value","NO","No","kept as actual; outside VAT base"],
 ["Municipal fees DT/DF (Dimotika Teli/Foros)","Cleaning & lighting municipal charge","Municipality (collected via bill)","m2 of property × municipal rate × days","NO – property-based","No (unless m2/municipality data wrong)","kept as actual; outside VAT base"],
 ["TAP (Telos Akinitis Periousias)","Property tax 0.025-0.035% of zone value","Municipality","m2 × zone value × age × days","NO – property-based","No","kept as actual"],
 ["ERT levy","Public broadcaster fee","ERT (via bill)","36 €/year prorated by days","NO","No","kept as actual; outside VAT base"],
 ["VAT (FPA)","Value added tax on electricity","State","6% of (A + B + EFK)","Indirectly – follows A","Only via lowering A","recomputed = 6% × (newA + B + EFK)"],
 ["Government subsidy (Kratiki Epidotisi T.E.M.)","Temporary state subsidy on energy when active","State","kWh when triggered","Applies to all","n/a","0 in these bills (none active)"],
]
r = hr + 1
for row in fee_rows:
    for i, v in enumerate(row, 1):
        cell = ws.cell(row=r, column=i, value=v); cell.border = BORDER; cell.alignment = WRAP
        if i == 5: cell.fill = WARN if v.startswith("YES") else GREY
    r += 1
autosize(ws, {"A":28,"B":34,"C":22,"D":28,"E":22,"F":26,"G":40})
ws.freeze_panes = "A4"

# =====================================================================
# SHEET 5: Market_Offers
# =====================================================================
ws = wb.create_sheet("Market_Offers")
ws["A1"] = "Market_Offers — researched current Greek offers (see source URLs & dates)"
ws["A1"].font = TITLE
cols = ["Offer ID","Provider","Product","Color/Category","Type","Segment","Energy €/kWh","Night €/kWh",
        "Adj. clause?","Paygio €/mo","Net disc % (qualified)","Disc conditions","Contract","Exit penalty",
        "High-consumption OK?","Source URL","Researched"]
hr = 3
for i, c in enumerate(cols, 1):
    ws.cell(row=hr, column=i, value=c)
style_header_row(ws, hr, len(cols))
r = hr + 1
offer_row = {}
for o in OFFERS:
    offer_row[o["id"]] = r
    vals = [o["id"], o["provider"], o["product"], o.get("color",""), o.get("type",""), o.get("segment",""),
            o["energy_rate"], o.get("night_rate"), "yes" if o.get("adj_clause") else "no",
            o["paygio"], o.get("net_disc_pct",0), o.get("disc_conditions",""), o.get("contract",""),
            o.get("exit_penalty",""), o.get("high_ok",""), o.get("source",""), o.get("date","")]
    for i, v in enumerate(vals, 1):
        cell = ws.cell(row=r, column=i, value=v); cell.border = BORDER; cell.alignment = WRAP
        if i in (7,8): cell.number_format = RATE
        if i == 10: cell.number_format = EUR
        if i == 11: cell.number_format = PCT
    r += 1
autosize(ws, {"A":9,"B":13,"C":24,"D":14,"E":10,"F":11,"G":10,"H":10,"I":10,"J":10,"K":12,
              "L":34,"M":12,"N":16,"O":16,"P":40,"Q":11})
ws.freeze_panes = "C4"
MO_SHEET = "Market_Offers"

# helper: excel cell refs into Market_Offers for an offer field
MO_COL = {"energy":"G","night":"H","paygio":"J","disc":"K"}
def mo_ref(offer_id, field):
    return f"'{MO_SHEET}'!{MO_COL[field]}{offer_row[offer_id]}"

# =====================================================================
# SHEET 6: Scenario_Model  (each offer applied to each supply, annualised)
# =====================================================================
ws = wb.create_sheet("Scenario_Model")
ws["A1"] = "Scenario_Model — estimated ANNUAL competitive-supply cost (component A only) of each offer on each supply"
ws["A1"].font = TITLE
ws["A2"] = ("Shows only the provider-dependent part (A = paygio + energy - discounts) annualised on each supply's sampled consumption. "
            "Regulated/municipal/ERT/taxes are identical across providers and are excluded here (they are added back in Backtest). "
            "Night-tariff supply S4 priced on TOTAL kWh at the offer's day rate unless the offer has a night rate (conservative).")
ws["A2"].alignment = WRAP; ws["A2"].font = Font(italic=True, size=9)
ws.merge_cells("A2:H2")
# per-supply annual kWh and day fraction
supply_ann = {}
for s in SUPPLIES:
    sb = [b for b in BILLS if b["supply_id"] == s["id"]]
    tot_kwh = sum(b["kwh_total"] for b in sb); tot_days = sum(b["days"] for b in sb)
    ann_kwh = tot_kwh/tot_days*365
    ann_night = sum(b["kwh_night"] for b in sb)/tot_days*365
    supply_ann[s["id"]] = (ann_kwh, ann_night, s)
hr = 4
head = ["Offer ID","Provider","Product"] + [f"{s['id']} {s['provider'][:4]} ann.A €" for s in SUPPLIES]
for i, c in enumerate(head, 1):
    ws.cell(row=hr, column=i, value=c)
style_header_row(ws, hr, len(head))
r = hr + 1
scen_first = r
for o in OFFERS:
    ws.cell(row=r, column=1, value=o["id"]).border = BORDER
    ws.cell(row=r, column=2, value=o["provider"]).border = BORDER
    ws.cell(row=r, column=3, value=o["product"]).border = BORDER
    for j, s in enumerate(SUPPLIES):
        ann_kwh, ann_night, sup = supply_ann[s["id"]]
        col = 4 + j
        oid = o["id"]
        seg_ok = (o.get("segment","any") in ("any", sup["seg"] if "seg" in sup else "")) or True
        # annual A = paygio*12 + energy*(1-disc)*kwh  (night portion at night rate if offer has one)
        night_rate_ref = mo_ref(oid,"night")
        # build formula: if offer night rate >0 use day on (ann-night)+night on night, else day on all
        f = (f"={mo_ref(oid,'paygio')}*12 + "
             f"IF({night_rate_ref}>0,"
             f"({ann_kwh-ann_night:.1f}*{mo_ref(oid,'energy')}+{ann_night:.1f}*{night_rate_ref}),"
             f"{ann_kwh:.1f}*{mo_ref(oid,'energy')})*(1-{mo_ref(oid,'disc')})")
        cell = ws.cell(row=r, column=col, value=f); cell.number_format = EUR; cell.border = BORDER
    r += 1
autosize(ws, {"A":9,"B":13,"C":24,"D":16,"E":16,"F":16,"G":16})
ws.freeze_panes = "D5"

# =====================================================================
# SHEET 7: Backtest  (apply recommended offer to each actual bill)
# =====================================================================
ws = wb.create_sheet("Backtest")
ws["A1"] = "Backtest — what each ACTUAL bill would have cost under the recommended alternative"
ws["A1"].font = TITLE
ws["A2"] = ("Method: replace component A (supplier paygio+energy-discounts) with the recommended offer; keep B (regulated), "
            "municipal, ERT and taxes EXACTLY as billed; recompute VAT = 6% × (newA + B + EFK). 5d2516a2 uses gross "
            "reconstruction (its payable is net of a 224.95€ on-account charge). Change the offer rates in Market_Offers and this recalculates.")
ws["A2"].alignment = WRAP; ws["A2"].font = Font(italic=True, size=9)
ws.merge_cells("A2:N2")
cols = ["File","Supply","Rec. offer","Days","kWh","Actual A €","Fixed non-A (B+tax+muni+ERT) €","Actual VAT €",
        "Actual total €","New A €","New VAT €","New total €","Saving €","Saving %"]
hr = 4
for i, c in enumerate(cols, 1):
    ws.cell(row=hr, column=i, value=c)
style_header_row(ws, hr, len(cols))
r = hr + 1
bt_first = r
for b in BILLS:
    sid = b["supply_id"]
    oid = REC[sid]
    o = next(x for x in OFFERS if x["id"] == oid)
    # gross handling for enanti bill
    if b["file"] == "5d2516a2":
        A_act = b["gross_A_supply"]; vat_act = b["gross_vat"]; total_act = b["gross_total"]
        vat_base_exclA = b["gross_vat_base"] - b["gross_A_supply"]
    else:
        A_act = b["A_supply"]; vat_act = b["vat"]; total_act = b["current_total"]
        vat_base_exclA = b["vat_base"] - b["A_supply"]
    fixed_nonA = round(total_act - A_act - vat_act, 2)
    days = b["days"]; kwh = b["kwh_total"]; night = b["kwh_night"]
    ws.cell(row=r, column=1, value=b["file"]).border = BORDER
    ws.cell(row=r, column=2, value=sid).border = BORDER
    ws.cell(row=r, column=3, value=oid).border = BORDER
    ws.cell(row=r, column=4, value=days).border = BORDER
    ws.cell(row=r, column=5, value=kwh).border = BORDER
    for col,val in [(6,A_act),(7,fixed_nonA),(8,vat_act),(9,total_act)]:
        c=ws.cell(row=r, column=col, value=round(val,2)); c.number_format=EUR; c.border=BORDER
    # New A formula referencing Market_Offers
    nr = mo_ref(oid,"night")
    newA = (f"={mo_ref(oid,'paygio')}*{days}/30 + "
            f"IF({nr}>0,({kwh-night}*{mo_ref(oid,'energy')}+{night}*{nr}),{kwh}*{mo_ref(oid,'energy')})"
            f"*(1-{mo_ref(oid,'disc')})")
    ca=ws.cell(row=r, column=10, value=newA); ca.number_format=EUR; ca.border=BORDER
    # New VAT = 6% * (newA + vat_base_exclA)
    cv=ws.cell(row=r, column=11, value=f"=0.06*(J{r}+{vat_base_exclA:.2f})"); cv.number_format=EUR; cv.border=BORDER
    # New total = fixed_nonA + newA + newVAT
    ct=ws.cell(row=r, column=12, value=f"={fixed_nonA:.2f}+J{r}+K{r}"); ct.number_format=EUR; ct.border=BORDER
    cs=ws.cell(row=r, column=13, value=f"=I{r}-L{r}"); cs.number_format=EUR; cs.border=BORDER
    cp=ws.cell(row=r, column=14, value=f"=IF(I{r}=0,0,M{r}/I{r})"); cp.number_format=PCT; cp.border=BORDER
    r += 1
# totals row
ws.cell(row=r, column=1, value="TOTAL").font = BOLD
for col in (6,7,8,9,10,11,12,13):
    L=get_column_letter(col)
    c=ws.cell(row=r, column=col, value=f"=SUM({L}{bt_first}:{L}{r-1})"); c.font=BOLD; c.number_format=EUR; c.fill=GREY
c=ws.cell(row=r, column=14, value=f"=M{r}/I{r}"); c.font=BOLD; c.number_format=PCT; c.fill=GREY
bt_total_row = r
ws.cell(row=r, column=5, value=f"=SUM(E{bt_first}:E{r-1})").font=BOLD
autosize(ws, {"A":11,"B":7,"C":10,"D":6,"E":7,"F":11,"G":22,"H":11,"I":13,"J":11,"K":11,"L":13,"M":11,"N":10})
ws.freeze_panes = "C5"

# per-supply summary
r += 2
ws.cell(row=r, column=1, value="Per-supply summary (current-period samples)").font = SUBHDR
r += 1
for i,c in enumerate(["Supply","Provider now","Rec. offer","Actual total €","New total €","Saving €","Saving %"],1):
    ws.cell(row=r, column=i, value=c)
style_header_row(ws, r, 7)
r += 1
for s in SUPPLIES:
    rows = [idx for idx,b in enumerate(BILLS) if b["supply_id"]==s["id"]]
    excel_rows = [bt_first+idx for idx in rows]
    act = "+".join(f"I{x}" for x in excel_rows)
    new = "+".join(f"L{x}" for x in excel_rows)
    ws.cell(row=r, column=1, value=s["id"]).border=BORDER
    ws.cell(row=r, column=2, value=s["provider"]).border=BORDER
    ws.cell(row=r, column=3, value=REC[s["id"]]).border=BORDER
    ca=ws.cell(row=r, column=4, value=f"={act}"); ca.number_format=EUR; ca.border=BORDER
    cn=ws.cell(row=r, column=5, value=f"={new}"); cn.number_format=EUR; cn.border=BORDER
    cs=ws.cell(row=r, column=6, value=f"=D{r}-E{r}"); cs.number_format=EUR; cs.border=BORDER
    cp=ws.cell(row=r, column=7, value=f"=F{r}/D{r}"); cp.number_format=PCT; cp.border=BORDER
    if s["priority"]=="HIGH":
        for cc in range(1,8): ws.cell(row=r, column=cc).fill = WARN
    r += 1

# =====================================================================
# =====================================================================
# SHEET: Tier_Ranking — rank WITHIN commitment tiers (strategy before price)
# =====================================================================
ws = wb.create_sheet("Tier_Ranking")
ws["A1"] = "Tier_Ranking — decide commit-to-a-price vs follow-the-market FIRST, then compare within the tier"
ws["A1"].font = TITLE
ws["A2"] = ("Floating, 1-year fixed and multi-year fixed are different RISK products, not one price line. Per supply, "
            "offers are grouped by commitment tier and ranked cheapest-first WITHIN each tier (computed on the sampled "
            "bills; re-run the builder after editing Market_Offers). Weigh the fixed term: a long lock in a volatile "
            "market is opportunity risk + an exit fee, not free certainty. Cheapest in each tier is highlighted.")
ws["A2"].alignment = WRAP; ws["A2"].font = Font(italic=True, size=9)
ws.merge_cells("A2:H2")

_VR = META.get("vat_rate", 0.06)
def _tr_econ(b):
    if b.get("enanti_credit") and b.get("gross_total") is not None:
        return b["gross_A_supply"], b["gross_vat"], b["gross_total"], b["gross_vat_base"] - b["gross_A_supply"]
    return b["A_supply"], b["vat"], b["current_total"], b["vat_base"] - b["A_supply"]
def _tr_newtotal(b, o):
    A, v, t, xb = _tr_econ(b); fx = t - A - v
    rd = o["energy_rate"]; rn = o.get("night_rate")
    e = b["kwh_day"]*rd + b["kwh_night"]*rn if (rn and b.get("kwh_night", 0) > 0) else b["kwh_total"]*rd
    nA = o["paygio"]*b["days"]/30.0 + e
    return fx + nA + _VR*(nA + xb)
def _tr_tier(o):
    is_fixed = o.get("color") == "Blue" or "fixed" in str(o.get("type", "")) or not o.get("adj_clause", True)
    if not is_fixed: return (0, "Floating — follow the market")
    m = 0
    for tok in str(o.get("contract", "")).replace("-", " ").split():
        if tok.isdigit(): m = int(tok); break
    if m == 0: return (1, "Fixed — term unspecified")
    if m <= 12: return (2, "Fixed ≤ 12 months")
    if m <= 24: return (3, "Fixed 13–24 months")
    return (4, "Fixed > 24 months")
def _tr_business(sup):
    return "business" in str(sup.get("customer_type", "")).lower() or sup.get("efk_per_kwh") == 0.005
def _tr_applicable(o, sup):
    seg = o.get("segment", "")
    if _tr_business(sup): return seg == "business"
    if "dual" in str(sup.get("register", "")): return seg in ("resi", "resi-night")
    return seg == "resi"

trow = 4
_hdr = ["Offer", "Provider", "Color", "Term", "Eff. €/kWh", "New total (sample) €", "Saving €", "Saving %"]
_NC = len(_hdr)
for sup in SUPPLIES:
    sb = [b for b in BILLS if b["supply_id"] == sup["id"]]
    actual = sum(_tr_econ(b)[2] for b in sb)
    hc = ws.cell(row=trow, column=1, value=f"{sup['id']} — {sup['property']}  ·  current actual (sample): €{actual:,.2f}")
    hc.font = SUBHDR; ws.merge_cells(start_row=trow, start_column=1, end_row=trow, end_column=_NC); trow += 1
    for i, h in enumerate(_hdr, 1): ws.cell(row=trow, column=i, value=h)
    style_header_row(ws, trow, _NC); trow += 1
    groups = {}
    for o in OFFERS:
        if not _tr_applicable(o, sup): continue
        k = _tr_tier(o); tot = sum(_tr_newtotal(b, o) for b in sb)
        groups.setdefault(k, []).append((tot, o))
    if not groups:
        ws.cell(row=trow, column=1, value="(no applicable offers for this segment/meter)"); trow += 1
    for k in sorted(groups):
        gc = ws.cell(row=trow, column=1, value=k[1]); gc.font = BOLD
        ws.merge_cells(start_row=trow, start_column=1, end_row=trow, end_column=_NC)
        for cc in range(1, _NC + 1): ws.cell(row=trow, column=cc).fill = GREY
        trow += 1
        for j, (tot, o) in enumerate(sorted(groups[k], key=lambda x: x[0])):
            sav = actual - tot
            vals = [o["product"], o["provider"], o.get("color", ""), o.get("contract", ""),
                    o["energy_rate"], round(tot, 2), round(sav, 2), (sav/actual if actual else 0)]
            for i, v in enumerate(vals, 1):
                cell = ws.cell(row=trow, column=i, value=v); cell.border = BORDER; cell.alignment = WRAP
                if i == 5: cell.number_format = RATE
                if i in (6, 7): cell.number_format = EUR
                if i == 8: cell.number_format = PCT
            if j == 0:
                for cc in range(1, _NC + 1): ws.cell(row=trow, column=cc).fill = GOOD
            trow += 1
    trow += 1
autosize(ws, {"A": 32, "B": 18, "C": 8, "D": 18, "E": 11, "F": 18, "G": 11, "H": 10})
ws.freeze_panes = "A4"


# SHEET 8: Recommendation
# =====================================================================
ws = wb.create_sheet("Recommendation")
ws["A1"] = "Recommendation — best move per supply"
ws["A1"].font = TITLE
cols = ["Supply","Property","Priority","Provider now → product","Recommended → product","Why","Est. monthly saving €","Action"]
hr = 3
for i, c in enumerate(cols, 1):
    ws.cell(row=hr, column=i, value=c)
style_header_row(ws, hr, len(cols))
r = hr + 1
for s in SUPPLIES:
    oid = REC[s["id"]]; o = next(x for x in OFFERS if x["id"]==oid)
    rec_text = offers["rec_text"].get(s["id"], {})
    vals = [s["id"], s["property"], s["priority"], f"{s['provider']} → {s['product']}",
            f"{o['provider']} → {o['product']}", rec_text.get("why",""), "", rec_text.get("action","")]
    for i, v in enumerate(vals, 1):
        cell = ws.cell(row=r, column=i, value=v); cell.border = BORDER; cell.alignment = WRAP
    if s["priority"]=="HIGH":
        ws.cell(row=r, column=3).fill = WARN; ws.cell(row=r, column=3).font = BOLD
    r += 1
autosize(ws, {"A":7,"B":32,"C":9,"D":26,"E":28,"F":50,"G":16,"H":40})
ws.freeze_panes = "A4"
r += 2
ws.cell(row=r, column=1, value="Overall expected savings: see Backtest TOTAL row. Headline figures are summarised in the report.").font = Font(italic=True)

out = os.path.join(HERE, "example_model.xlsx")
wb.save(out)
print("Saved:", out)
print("Offers:", len(OFFERS), "| Recommendations:", REC)
